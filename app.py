import streamlit as st
import pandas as pd
import io
import random
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Agenda Bersih-Bersih ELCC", layout="wide")

st.title("🧹 Sistem Pembagian Agenda Bersih-Bersih ELCC 2026")
st.write("Membagi 37 Anggota dari 4 Divisi secara merata ke dalam 7 bulan (Juni - Desember) dengan format output vertikal berkala.")

# ----------------- SIDEBAR INPUT -----------------
st.sidebar.header("📁 Unggah Data Anggota")
uploaded_file = st.sidebar.file_uploader(
    "Unggah File Excel (Harus ada kolom: Nomor, Nama, NBI, Divisi)", 
    type=["xlsx", "xls"],
    key="data_anggota"
)

st.sidebar.markdown("---")
st.sidebar.subheader("⚙️ Pengaturan Pengacakan")
use_seed = st.sidebar.checkbox("Kunci Hasil Pengacakan (Lock Seed)", value=True)
seed_value = st.sidebar.number_input("Nilai Seed", min_value=1, max_value=9999, value=42, disabled=not use_seed)

# ----------------- LOGIKA UTAMA -----------------
if st.sidebar.button("🎲 Susun Agenda Bersih-Bersih", type="primary"):
    if not uploaded_file:
        st.error("❌ Mohon unggah file Excel data anggota terlebih dahulu!")
    else:
        with st.spinner("Sedang memproses pembagian divisi dan menyusun layout..."):
            try:
                # 1. Baca Data
                df_raw = pd.read_excel(uploaded_file)
                df_raw.columns = df_raw.columns.str.strip()
                
                # Validasi Kolom
                required_cols = ['Nomor', 'Nama', 'NBI', 'Divisi']
                for col in required_cols:
                    if col not in df_raw.columns:
                        st.error(f"❌ Kolom '{col}' tidak ditemukan di file Excel Anda!")
                        st.stop()
                
                df_raw['Nama'] = df_raw['Nama'].astype(str).str.strip().str.title()
                df_raw['NBI'] = df_raw['NBI'].astype(str).str.strip()
                df_raw['Divisi'] = df_raw['Divisi'].astype(str).str.strip()
                
                # Atur Seed
                if use_seed:
                    random.seed(seed_value)
                else:
                    random.seed(None)
                
                # 2. Algoritma Pemerataan Divisi (Stratified Round-Robin)
                divisi_groups = {}
                for div_name, group in df_raw.groupby('Divisi'):
                    records = group[['Nama', 'NBI', 'Divisi']].to_dict('records')
                    random.shuffle(records)
                    divisi_groups[div_name] = records
                
                antrean_merata = []
                while any(divisi_groups.values()):
                    for div_name in list(divisi_groups.keys()):
                        if divisi_groups[div_name]:
                            antrean_merata.append(divisi_groups[div_name].pop(0))
                
                # Struktur Kuota Tetap
                daftar_bulan = [
                    {"nama": "JUNI 2026", "kuota": 6},
                    {"nama": "JULI 2026", "kuota": 5},
                    {"nama": "AGUSTUS 2026", "kuota": 5},
                    {"nama": "SEPTEMBER 2026", "kuota": 5},
                    {"nama": "OKTOBER 2026", "kuota": 5},
                    {"nama": "NOVEMBER 2026", "kuota": 6},
                    {"nama": "DESEMBER 2026", "kuota": 5}
                ]
                
                # Pecah antrean ke dalam struktur dictionary per bulan
                distribusi_bulan = {}
                idx_antrean = 0
                for bulan in daftar_bulan:
                    list_anggota_bulan = []
                    for no_urut in range(1, bulan["kuota"] + 1):
                        if idx_antrean < len(antrean_merata):
                            list_anggota_bulan.append(antrean_merata[idx_antrean])
                            idx_antrean += 1
                    distribusi_bulan[bulan["nama"]] = list_anggota_bulan

                # ----------------- VISUALISASI DASHBOARD -----------------
                st.success("🎉 Agenda bersih-bersih berhasil disusun secara merata!")
                
                # Tampilkan Preview Tabs di Web
                tabs_bulan = st.tabs([b["nama"] for b in daftar_bulan])
                for i, bulan in enumerate(daftar_bulan):
                    with tabs_bulan[i]:
                        df_view = pd.DataFrame(distribusi_bulan[bulan["nama"]])
                        df_view.insert(0, 'No.', range(1, len(df_view) + 1))
                        st.dataframe(df_view, use_container_width=True)

                # --- 🛠️ EXCEL GENERATION DENGAN OPENPYXL (LAY OUT VERTIKAL KETAT) 🛠️ ---
                wb = Workbook()
                ws = wb.active
                ws.title = "Agenda Bersih-Bersih"
                ws.views.sheetView[0].showGridLines = True
                
                # Setup Style Komponen
                font_judul = Font(name="Calibri", size=14, bold=True)
                font_header_bulan = Font(name="Calibri", size=12, bold=True, color="154360")
                font_th = Font(name="Calibri", size=11, bold=True, color="000000")
                font_td = Font(name="Calibri", size=11)
                
                center_align = Alignment(horizontal="center", vertical="center")
                left_align_title = Alignment(horizontal="center", vertical="center")
                
                fill_hijau = PatternFill(start_color="2ECC71", end_color="2ECC71", fill_type="solid") # Hijau bersih cerah
                
                thin_side = Side(border_style="thin", color="000000")
                border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
                
                # Kolom B sampai E (karena kolom A wajib kosong)
                start_col = 2 
                columns_table = ["No.", "Nama", "NBI", "Divisi"]
                last_col_letter = get_column_letter(start_col + len(columns_table) - 1)
                
                # 1. Buat Judul Utama di Baris 3 & 4 (Merge & Center dari B sampai E)
                ws.merge_cells(f"B3:{last_col_letter}3")
                ws.merge_cells(f"B4:{last_col_letter}4")
                ws["B3"] = "AGENDA BERSIH - BERSIH"
                ws["B4"] = "SEKRETARIAT ELCC 2026"
                
                ws["B3"].font = font_judul
                ws["B3"].alignment = center_align
                ws["B4"].font = font_judul
                ws["B4"].alignment = center_alignment
                
                ws.row_dimensions[3].height = 22
                ws.row_dimensions[4].height = 22
                
                # 2. Iterasi Penyusunan Tabel Bulan ke Bawah
                # Tabel pertama dimulai di baris ke-7 (memberikan ruang setelah 5 baris pertama)
                current_row = 7 
                
                for bulan in daftar_bulan:
                    nama_bulan = bulan["nama"]
                    data_anggota = distribusi_bulan[nama_bulan]
                    
                    # A. Cetak Keterangan Bulan (Misal: JUNI 2026)
                    ws.cell(row=current_row, column=start_col, value=nama_bulan).font = font_header_bulan
                    ws.row_dimensions[current_row].height = 20
                    current_row += 1
                    
                    # B. Cetak Judul Kolom Tabel (No., Nama, NBI, Divisi) -> Hijau Cerah & Center
                    ws.row_dimensions[current_row].height = 26
                    for col_idx, col_name in enumerate(columns_table, start=start_col):
                        cell = ws.cell(row=current_row, column=col_idx, value=col_name)
                        cell.font = font_th
                        cell.fill = fill_hijau
                        cell.alignment = center_align
                        cell.border = border_cell
                    current_row += 1
                    
                    # C. Cetak Isi Data Anggota Bulan Ini -> Center & Border
                    for urutan, anggota in enumerate(data_anggota, start=1):
                        ws.row_dimensions[current_row].height = 20
                        
                        # Ambil nilai baris
                        vals = [urutan, anggota["Nama"], anggota["NBI"], anggota["Divisi"]]
                        
                        for col_idx, val in enumerate(vals, start=start_col):
                            cell = ws.cell(row=current_row, column=col_idx, value=val)
                            cell.font = font_td
                            cell.alignment = center_align
                            cell.border = border_cell
                        current_row += 1
                    
                    # D. Berikan Jarak Jeda Tepat 3 Baris Kosong untuk Tabel Bulan Berikutnya
                    current_row += 3 
                
                # 3. Pengaturan Lebar Kolom (Kolom A Sengaja Dikosongkan)
                ws.column_dimensions['A'].width = 4
                for col_idx in range(start_col, start_col + len(columns_table)):
                    col_letter = get_column_letter(col_idx)
                    # Deteksi lebar maksimal secara dinamis agar tidak terpotong
                    max_len = 0
                    for r in range(6, current_row):
                        val_str = str(ws.cell(row=r, column=col_idx).value or '')
                        if len(val_str) > max_len:
                            max_len = len(val_str)
                    ws.column_dimensions[col_letter].width = max(max_len + 5, 12)
                
                # Siapkan file blob untuk di-download
                output = io.BytesIO()
                wb.save(output)
                processed_data = output.getvalue()
                
                st.sidebar.markdown("---")
                st.sidebar.success("💾 Format Dokumen Vertikal Siap!")
                st.sidebar.download_button(
                    label="📥 Download Agenda Terformat (.xlsx)",
                    data=processed_data,
                    file_name="Agenda_Bersih_ELCC_2026.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
            except Exception as e:
                st.error(f"Terjadi kesalahan teknis penyusunan file: {str(e)}")
